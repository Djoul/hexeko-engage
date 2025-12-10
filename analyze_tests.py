#!/usr/bin/env python3
"""
Script d'analyse des résultats de tests PHPUnit
Génère un fichier Excel et un rapport Markdown
"""

import re
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def find_latest_test_file():
    """Trouve le fichier output-test-*.txt le plus récent"""
    files = list(Path('.').glob('output-test-*.txt'))
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def parse_test_output(content):
    """Parse le contenu du fichier de tests PHPUnit"""
    data = defaultdict(lambda: {'errors': [], 'failures': []})

    # Pattern pour capturer les erreurs et failures
    error_pattern = r'(\d+)\)\s+(Tests\\[^:]+)::(\w+)\s*\n([^\n]+)'

    # Trouver la section des erreurs
    errors_match = re.search(r'There (?:was|were) \d+ errors?:', content, re.IGNORECASE)
    if errors_match:
        start = errors_match.end()
        end_match = re.search(r'(FAILURES!|There (?:was|were) \d+ failures?:|Tests: \d+)', content[start:], re.IGNORECASE)
        end = start + end_match.start() if end_match else len(content)
        errors_section = content[start:end]

        for match in re.finditer(error_pattern, errors_section):
            num, class_name, method, message = match.groups()
            data[class_name]['errors'].append({
                'method': method,
                'message': message.strip()
            })

    # Trouver la section des failures
    failures_match = re.search(r'There (?:was|were) \d+ failures?:', content, re.IGNORECASE)
    if failures_match:
        start = failures_match.end()
        end_match = re.search(r'(FAILURES!|Tests: \d+)', content[start:], re.IGNORECASE)
        end = start + end_match.start() if end_match else len(content)
        failures_section = content[start:end]

        for match in re.finditer(error_pattern, failures_section):
            num, class_name, method, message = match.groups()
            data[class_name]['failures'].append({
                'method': method,
                'message': message.strip()
            })

    # Calculer les totaux
    for class_name in data:
        data[class_name]['total_errors'] = len(data[class_name]['errors'])
        data[class_name]['total_failures'] = len(data[class_name]['failures'])
        data[class_name]['total'] = data[class_name]['total_errors'] + data[class_name]['total_failures']

    return dict(data)


def get_priority(total):
    """Détermine la priorité et la couleur en fonction du nombre de problèmes"""
    if total >= 20:
        return '🔴 CRITIQUE', 'DC3545'
    elif total >= 10:
        return '🟠 HAUTE', 'FFC107'
    elif total >= 5:
        return '🟡 MOYENNE', 'FF9800'
    else:
        return '🟢 BASSE', '4CAF50'


def create_excel_report(data, timestamp):
    """Crée le rapport Excel avec 5 feuilles"""
    wb = Workbook()

    # Styles communs
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Trier les classes par nombre total de problèmes
    sorted_classes = sorted(data.items(), key=lambda x: x[1]['total'], reverse=True)

    # === FEUILLE 1: Synthèse par Classe ===
    ws1 = wb.active
    ws1.title = "Synthèse par Classe"
    headers1 = ['#', 'Classe de Test', 'Namespace Complet', 'Errors', 'Failures', 'Total', 'Priorité']
    ws1.append(headers1)

    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    for idx, (class_name, info) in enumerate(sorted_classes, 1):
        short_name = class_name.split('\\')[-1]
        priority_text, priority_color = get_priority(info['total'])

        row = [idx, short_name, class_name, info['total_errors'], info['total_failures'], info['total'], priority_text]
        ws1.append(row)

        for col_num in range(1, len(row) + 1):
            cell = ws1.cell(ws1.max_row, col_num)
            cell.border = border
            if col_num in [1, 4, 5, 6]:
                cell.alignment = center_alignment
            if col_num == 7:
                cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type='solid')

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 15

    # === FEUILLE 2: Détails des Errors ===
    ws2 = wb.create_sheet("Détails des Errors")
    headers2 = ['#', 'Classe', 'Méthode de Test', 'Type', "Message d'Erreur"]
    ws2.append(headers2)

    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    error_num = 1
    error_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    for class_name, info in sorted_classes:
        for error in info['errors']:
            row = [error_num, class_name.split('\\')[-1], error['method'], 'ERROR', error['message']]
            ws2.append(row)

            for col_num in range(1, len(row) + 1):
                cell = ws2.cell(ws2.max_row, col_num)
                cell.border = border
                if col_num in [1, 4]:
                    cell.alignment = center_alignment
                if col_num == 4:
                    cell.fill = error_fill

            error_num += 1

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 80

    # === FEUILLE 3: Détails des Failures ===
    ws3 = wb.create_sheet("Détails des Failures")
    headers3 = ['#', 'Classe', 'Méthode de Test', 'Type', "Message d'Échec"]
    ws3.append(headers3)

    for col_num, header in enumerate(headers3, 1):
        cell = ws3.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    failure_num = 1
    failure_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')
    for class_name, info in sorted_classes:
        for failure in info['failures']:
            row = [failure_num, class_name.split('\\')[-1], failure['method'], 'FAILURE', failure['message']]
            ws3.append(row)

            for col_num in range(1, len(row) + 1):
                cell = ws3.cell(ws3.max_row, col_num)
                cell.border = border
                if col_num in [1, 4]:
                    cell.alignment = center_alignment
                if col_num == 4:
                    cell.fill = failure_fill

            failure_num += 1

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 80

    # === FEUILLE 4: TODO List ===
    ws4 = wb.create_sheet("TODO List")
    headers4 = ['☐', 'Priorité', 'Classe Complète', 'Total Problèmes', 'Notes / Actions']
    ws4.append(headers4)

    for col_num, header in enumerate(headers4, 1):
        cell = ws4.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    for class_name, info in sorted_classes:
        priority_text, priority_color = get_priority(info['total'])
        notes = f"{info['total_errors']} errors, {info['total_failures']} failures"
        row = ['☐', priority_text, class_name, info['total'], notes]
        ws4.append(row)

        for col_num in range(1, len(row) + 1):
            cell = ws4.cell(ws4.max_row, col_num)
            cell.border = border
            if col_num in [1, 2, 4]:
                cell.alignment = center_alignment
            if col_num == 2:
                cell.fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type='solid')

    ws4.column_dimensions['A'].width = 5
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 50
    ws4.column_dimensions['D'].width = 15
    ws4.column_dimensions['E'].width = 40

    # === FEUILLE 5: Statistiques Globales ===
    ws5 = wb.create_sheet("Statistiques Globales")

    total_classes = len(data)
    total_errors = sum(info['total_errors'] for info in data.values())
    total_failures = sum(info['total_failures'] for info in data.values())
    total_problems = total_errors + total_failures

    critique = sum(1 for info in data.values() if info['total'] >= 20)
    haute = sum(1 for info in data.values() if 10 <= info['total'] < 20)
    moyenne = sum(1 for info in data.values() if 5 <= info['total'] < 10)
    basse = sum(1 for info in data.values() if info['total'] < 5)

    stats = [
        ['📊 RÉSUMÉ DES TESTS', ''],
        ['Total classes avec problèmes', total_classes],
        ['Total Errors', total_errors],
        ['Total Failures', total_failures],
        ['Total problèmes', total_problems],
        ['', ''],
        ['📈 RÉPARTITION PAR PRIORITÉ', ''],
        ['🔴 Critique (≥20)', critique],
        ['🟠 Haute (10-19)', haute],
        ['🟡 Moyenne (5-9)', moyenne],
        ['🟢 Basse (<5)', basse],
    ]

    for row in stats:
        ws5.append(row)

    for row_num in range(1, ws5.max_row + 1):
        cell_a = ws5.cell(row_num, 1)
        cell_b = ws5.cell(row_num, 2)

        if row_num in [1, 7]:
            cell_a.font = Font(bold=True, size=14, color='FFFFFF')
            cell_a.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        else:
            cell_a.font = Font(bold=True)
            cell_b.alignment = center_alignment

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 20

    filename = f'analyse_tests_phpunit_{timestamp}.xlsx'
    wb.save(filename)
    return filename


def create_markdown_report(data, timestamp):
    """Crée le rapport Markdown"""
    sorted_classes = sorted(data.items(), key=lambda x: x[1]['total'], reverse=True)

    total_classes = len(data)
    total_errors = sum(info['total_errors'] for info in data.values())
    total_failures = sum(info['total_failures'] for info in data.values())
    total_problems = total_errors + total_failures

    md = [
        "# 📋 Analyse des Tests PHPUnit - TODO List\n",
        "## 📊 Vue d'Ensemble\n",
        f"- **Total classes avec problèmes:** {total_classes}",
        f"- **Total Errors:** {total_errors}",
        f"- **Total Failures:** {total_failures}",
        f"- **Total problèmes:** {total_problems}\n",
        "## 🎯 Classes Prioritaires (par nombre de problèmes)\n"
    ]

    for idx, (class_name, info) in enumerate(sorted_classes, 1):
        priority_text, _ = get_priority(info['total'])
        short_name = class_name.split('\\')[-1]

        md.append(f"### {idx}. {priority_text.split()[0]} {short_name} - **{info['total']} problèmes**\n")
        md.append(f"- **Priorité:** {priority_text.split()[1]}")
        md.append(f"- **Namespace:** `{class_name}`")
        md.append(f"- **Errors:** {info['total_errors']}")
        md.append(f"- **Failures:** {info['total_failures']}\n")

    md.append("\n## 🔍 Patterns Identifiés\n")

    error_messages = []
    for info in data.values():
        for error in info['errors']:
            error_messages.append(error['message'])
        for failure in info['failures']:
            error_messages.append(failure['message'])

    common_patterns = Counter(error_messages).most_common(5)

    if common_patterns:
        md.append("### Messages d'erreur les plus fréquents:\n")
        for msg, count in common_patterns:
            if count > 1:
                md.append(f"- **{count}x:** {msg[:100]}{'...' if len(msg) > 100 else ''}")

    md.append("\n## 📝 Actions Recommandées\n")
    md.append("### 🔴 Priorité CRITIQUE (≥20 problèmes)")
    md.append("- Traiter immédiatement les classes critiques")
    md.append("- Isoler les tests défaillants et les corriger un par un\n")

    md.append("### 🟠 Priorité HAUTE (10-19 problèmes)")
    md.append("- Planifier la correction dans le sprint en cours")
    md.append("- Analyser les causes racines communes\n")

    md.append("### 🟡 Priorité MOYENNE (5-9 problèmes)")
    md.append("- Corriger progressivement lors du refactoring")
    md.append("- Améliorer la couverture de tests\n")

    md.append("### 🟢 Priorité BASSE (<5 problèmes)")
    md.append("- Traiter lors de la maintenance courante")
    md.append("- Optimiser les tests existants")

    filename = f'RAPPORT_ANALYSE_TESTS_{timestamp}.md'
    with open(filename, 'w', encoding='utf-8') as f:
        f.write('\n'.join(md))

    return filename


def main():
    """Fonction principale"""
    print("🔍 Recherche du fichier de tests le plus récent...")

    test_file = find_latest_test_file()
    if not test_file:
        print("❌ Aucun fichier output-test-*.txt trouvé à la racine du projet")
        sys.exit(1)

    print(f"✅ Fichier trouvé: {test_file}")
    print("📖 Parsing du fichier...")

    with open(test_file, 'r', encoding='utf-8') as f:
        content = f.read()

    data = parse_test_output(content)

    if not data:
        print("❌ Aucune erreur ou failure détectée dans le fichier")
        sys.exit(1)

    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')

    print("📊 Génération du rapport Excel...")
    excel_file = create_excel_report(data, timestamp)

    print("📝 Génération du rapport Markdown...")
    md_file = create_markdown_report(data, timestamp)

    total_classes = len(data)
    total_errors = sum(info['total_errors'] for info in data.values())
    total_failures = sum(info['total_failures'] for info in data.values())
    total_problems = total_errors + total_failures

    sorted_classes = sorted(data.items(), key=lambda x: x[1]['total'], reverse=True)

    print("\n✅ Analyse terminée !\n")
    print("📊 Statistiques:")
    print(f"   - {total_classes} classes avec problèmes")
    print(f"   - {total_errors} erreurs (Errors)")
    print(f"   - {total_failures} échecs (Failures)")
    print(f"   - {total_problems} problèmes au total\n")

    print("📦 Fichiers générés:")
    print(f"   - {excel_file}")
    print(f"   - {md_file}\n")

    print("🎯 Top 3 des classes critiques:")
    for idx, (class_name, info) in enumerate(sorted_classes[:3], 1):
        short_name = class_name.split('\\')[-1]
        print(f"   {idx}. {short_name} ({info['total']} problèmes)")


if __name__ == '__main__':
    main()
